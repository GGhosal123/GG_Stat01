# GG_Stat_Ref
from GG_Code_Sep_01 import *

def main():
    import os
    import pandas as pd
    import matplotlib.pyplot as plt
    import statistics
    import itertools
    import math
    from itertools import permutations
    from collections import Counter

    os.system('cls')
    #m_path= 'C:\Users\goura\OneDrive\Desktop\Stat_Ref_1\IGNOU-MSCAST\Stat_Ref_1.xlsx'
    # read by default 1st sheet of an excel file
    df = pd.read_excel('Stat_Ref_1.xlsx','Raw-Data1')
    df.fillna(value='')
   
    sep_char('=')
    print(df)
    sep_char('=')
    my_samples = int(input("Enter number of Samples to be taken : "))

    my_sal = pd.read_excel('Stat_Ref_1.xlsx','Raw-Data1',usecols=['Monthly_Salary'])
    m_mean = statistics.mean(my_sal['Monthly_Salary'])
    m_stddev = statistics.stdev(my_sal['Monthly_Salary'])

    perm = itertools.product(df['Employee'], repeat=my_samples)
    permX = itertools.product(df['Employee'], repeat=my_samples)
    df1 = pd.DataFrame(perm)
    perm1 = itertools.product(df['Monthly_Salary'], repeat=my_samples)
    permY = itertools.product(df['Monthly_Salary'], repeat=my_samples)
    df2 = pd.DataFrame(perm1)

    sep_char('=')
    cnt = 0
    for i in list(permX): 
        cnt += 1
        print (cnt,end=' ')
        print(i)

    sep_char('=')
    cnt = 0
    m_freq_list=[]

    for j in list(permY): 
        cnt +=1

        m_sample_mean =round(statistics.mean(j),2)
        
        print(cnt, end=' ')
        print(j, end=' ')
        print(m_sample_mean)

        m_freq_list.append(m_sample_mean)

    f_dict = dict(Counter(m_freq_list))
   
    print("Frequency-Dict of Sample Mean :",end= ' ')
    print(f_dict)
    print("Frequency-List of Sample Mean :")
    f_list = list(f_dict.items())
    f_list.sort()
    cnt1=0
    m_processed_list=[]
    for i in f_list:
        cnt1 +=1
        m_x_bar,m_frequency = i
        #print (cnt1,end=' ')
        #print (m_x_bar,end=' ')
        #print(m_frequency,end=' ')
        m_probability = round(m_frequency/cnt,2)
        #print(m_probability)
        m_processed_list.append([cnt1,m_x_bar,m_frequency,m_probability])

        df3 = pd.DataFrame(m_processed_list)

    print(m_processed_list)
    #print(f_list)

    from openpyxl import load_workbook

    wb = load_workbook('Stat_Ref_1.xlsx')
    if 'Processed_Data_1' in wb.sheetnames:
        wb.remove(wb['Processed_Data_1'])
    if 'Processed_Data_2' in wb.sheetnames:
        wb.remove(wb['Processed_Data_2'])
    if 'Processed_Data_3' in wb.sheetnames:
        wb.remove(wb['Processed_Data_3'])
    wb.save('Stat_Ref_1.xlsx')

    writer = pd.ExcelWriter('Stat_Ref_1.xlsx', engine='openpyxl', mode='a')
    df1.to_excel(writer, sheet_name='Processed_Data_1')
    df2.to_excel(writer, sheet_name='Processed_Data_2')
    df3.to_excel(writer, sheet_name='Processed_Data_3')
    
    #writer.save()
    writer.close()

    sep_char('=')
    print('Population Mean is :',end=' ')
    print(m_mean)
    print('Population Standard Deviation is :',end=' ')
    print(round(m_stddev,2))

    sep_char('=')
    m_tot=0
    for i in m_processed_list:
        cnt1,m_x_bar,m_frequency,m_probability = i
        m_tot = m_tot + (m_x_bar * m_frequency)

    m_sample_mean2 = (1/cnt)* m_tot
    print("Mean of Sample Means :" , m_sample_mean2)
    sep_char('=')

    for i in m_processed_list:
        cnt1,m_x_bar,m_frequency,m_probability = i
        print('m_x_bar : ',m_x_bar)
        print('m_sample_mean2 : ', m_sample_mean2)
        m_tot1 = m_frequency * (m_x_bar - m_sample_mean2)**2

    print(m_tot1)
    print(cnt)
    m_tot2= (1/cnt) * m_tot1
    print(m_tot2)
    print("Standard Deviation of the sampling distribution of the mean : ",end=' ' )
    
    print(math.sqrt(m_tot2))
    sep_char('=')

    p_title = "Sampling distribution of mean when sample size is " + str(my_samples)

    #plt.figure(figsize=(2,2))
    plt.title(p_title)
    plt.hist(df3,density= True)
    plt.show()


# Using the special variable 
# __name__
if __name__ == "__main__":
    main()